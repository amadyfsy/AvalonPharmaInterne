import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def generate_excel(headers, data, sheet_title="Export"):
    """
    Generate an Excel file from a list of headers and a list of data rows.
    Returns a BytesIO object suitable for Flask send_file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title
    
    # Headers styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Write headers
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header_title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)
            
    # Auto-adjust column widths (basic approximation)
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
            
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
