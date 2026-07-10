"""Export Excel / PDF des statistiques."""
from __future__ import annotations

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .statistiques_queries import PeriodeStats


def _header_style():
    fill = PatternFill(start_color='0066CC', end_color='0066CC', fill_type='solid')
    font = Font(color='FFFFFF', bold=True)
    return fill, font


def _write_sheet(ws, headers, rows):
    fill, font = _header_style()
    for col, title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal='center')
    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    for col in ws.columns:
        width = max(len(str(c.value or '')) for c in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(width, 48)


def generate_statistiques_excel(bundle: dict) -> io.BytesIO:
    periode: PeriodeStats = bundle['periode']
    kpi = bundle['kpi']
    charts = bundle['chart_data']
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Synthèse'
    _write_sheet(
        ws,
        ['Indicateur', 'Valeur'],
        [
            ['Période', periode.label],
            ['Du', periode.debut.strftime('%d/%m/%Y')],
            ['Au', periode.fin.strftime('%d/%m/%Y')],
            ['CA TTC', kpi['ca_ttc']],
            ['CA HT', kpi['ca_ht']],
            ['Variation CA', f"{kpi['ca_variation']:.1f} %"],
            ['Marge brute', kpi['marge_brute']],
            ['Taux marge', f"{kpi['taux_marge']:.1f} %"],
            ['Achats TTC', kpi['achats_ttc']],
            ['Dépenses validées', kpi['depenses_ttc']],
            ['Bénéfice net estimé', kpi['benefice_net']],
            ['Créances clients', kpi['creances_total']],
            ['Valeur stock', kpi['valeur_stock']],
            ['Factures', kpi['nb_factures']],
            ['Commandes fournisseur', kpi['nb_commandes']],
        ],
    )

    evo = charts.get('evolution') or {}
    ws2 = wb.create_sheet('Évolution')
    rows = []
    for i, lbl in enumerate(evo.get('labels') or []):
        rows.append([
            lbl,
            (evo.get('ca') or [0])[i] if i < len(evo.get('ca') or []) else 0,
            (evo.get('achats') or [0])[i] if i < len(evo.get('achats') or []) else 0,
            (evo.get('depenses') or [0])[i] if i < len(evo.get('depenses') or []) else 0,
            (evo.get('marge') or [0])[i] if i < len(evo.get('marge') or []) else 0,
        ])
    _write_sheet(ws2, ['Période', 'CA TTC', 'Achats', 'Dépenses', 'Marge brute'], rows)

    ws3 = wb.create_sheet('Top clients')
    _write_sheet(
        ws3,
        ['Client', 'Type', 'Nb factures', 'CA TTC'],
        [[c['nom'], c['type'], c['nb_factures'], c['montant']] for c in bundle.get('top_clients') or []],
    )

    top_p = charts.get('top_produits') or {}
    ws4 = wb.create_sheet('Top produits')
    rows_p = []
    for i, lbl in enumerate(top_p.get('labels') or []):
        rows_p.append([
            lbl,
            (top_p.get('qty') or [0])[i] if i < len(top_p.get('qty') or []) else 0,
            (top_p.get('montants') or [0])[i] if i < len(top_p.get('montants') or []) else 0,
        ])
    _write_sheet(ws4, ['Produit', 'Quantité', 'Montant HT'], rows_p)

    dep = charts.get('depenses_categorie') or {}
    ws5 = wb.create_sheet('Dépenses')
    _write_sheet(
        ws5,
        ['Catégorie', 'Montant TTC'],
        [[lbl, dep['data'][i]] for i, lbl in enumerate(dep.get('labels') or [])],
    )

    cre = charts.get('creances') or {}
    ws6 = wb.create_sheet('Créances')
    _write_sheet(
        ws6,
        ['Tranche', 'Montant', 'Nb factures'],
        [
            [cre['labels'][i], cre['data'][i], cre['counts'][i]]
            for i in range(len(cre.get('labels') or []))
        ],
    )

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def export_filename(periode: PeriodeStats, ext: str) -> str:
    slug = periode.label.replace(' ', '_').replace('/', '-')
    ts = datetime.now().strftime('%Y%m%d')
    return f'Statistiques_{slug}_{ts}.{ext}'
